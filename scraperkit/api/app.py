"""
FastAPI application — dashboard API, job runner, config manager, and SSE log streaming.

Start with:
    scraperkit serve
or:
    uvicorn scraperkit.api.app:app --reload
"""
from __future__ import annotations

import asyncio
import os
from pathlib import Path
from typing import AsyncGenerator

from fastapi import FastAPI, HTTPException, Body
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from scraperkit.logging.db import RunStore


# ------------------------------------------------------------------ request models

class StartJobRequest(BaseModel):
    config_path: str | None = None   # path to an existing .yaml file
    config_name: str | None = None   # name of a managed config (in configs/)
    inline_config: str | None = None # raw YAML content


class WriteConfigRequest(BaseModel):
    name: str
    content: str


class ValidateConfigRequest(BaseModel):
    content: str


# ------------------------------------------------------------------ app factory

def create_app(
    db_path: str | Path = "scraperkit.db",
    configs_dir: str | Path = "configs",
) -> FastAPI:
    app = FastAPI(
        title="ScraperKit",
        description="Config-driven scraping and automation platform",
        version="0.1.0",
    )

    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_methods=["*"],
        allow_headers=["*"],
    )

    store = RunStore(db_path)

    from scraperkit.api.jobs import get_manager
    from scraperkit.api.configs import get_config_manager

    jobs = get_manager()
    cfgs = get_config_manager(str(configs_dir))

    # ---------------------------------------------------------------- run history

    @app.get("/api/runs")
    def list_runs(project: str | None = None, limit: int = 50):
        return store.list_runs(project=project, limit=limit)

    @app.get("/api/runs/{run_id}")
    def get_run(run_id: str):
        run = store.get_run(run_id)
        if not run:
            raise HTTPException(status_code=404, detail=f"Run '{run_id}' not found")
        return run

    @app.get("/api/projects/{project}/latest")
    def get_latest_run(project: str):
        run = store.get_latest_run(project)
        if not run:
            raise HTTPException(status_code=404, detail=f"No runs found for '{project}'")
        return run

    @app.get("/api/projects/{project}/stats")
    def get_project_stats(project: str):
        return store.project_stats(project)

    @app.get("/api/projects")
    def list_projects():
        # Merge DB projects + configs directory
        runs = store.list_runs(limit=1000)
        seen: dict[str, dict] = {}
        for run in runs:
            p = run["project"]
            if p not in seen:
                seen[p] = {
                    "project": p,
                    "total_runs": 0,
                    "last_run": run["started_at"],
                    "last_status": run["status"],
                }
            seen[p]["total_runs"] += 1

        # Add projects that have a config but no runs yet
        for cfg in cfgs.list_configs():
            if cfg["name"] not in seen:
                seen[cfg["name"]] = {
                    "project": cfg["name"],
                    "total_runs": 0,
                    "last_run": None,
                    "last_status": None,
                }
        return list(seen.values())

    # ---------------------------------------------------------------- live jobs

    @app.get("/api/jobs")
    def list_jobs():
        return jobs.list_jobs()

    @app.get("/api/jobs/{run_id}")
    def get_job(run_id: str):
        job = jobs.get_job(run_id)
        if not job:
            raise HTTPException(status_code=404, detail=f"Job '{run_id}' not found")
        from scraperkit.api.jobs import JobManager
        return JobManager._job_to_dict(job)

    @app.post("/api/jobs", status_code=201)
    def start_job(req: StartJobRequest):
        """Start a crawl job. Accepts a config_path, config_name, or inline_config."""
        import tempfile

        tmp_path: str | None = None

        if req.inline_config:
            # Write inline YAML to a temp file
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".yaml", delete=False, encoding="utf-8"
            ) as f:
                f.write(req.inline_config)
                tmp_path = f.name
            config_path = tmp_path
            # Extract project name from YAML
            import yaml
            try:
                data = yaml.safe_load(req.inline_config)
                project = data.get("name", "inline")
            except Exception:
                project = "inline"

        elif req.config_name:
            try:
                content = cfgs.read_config(req.config_name)
            except FileNotFoundError:
                raise HTTPException(status_code=404, detail=f"Config '{req.config_name}' not found")
            with tempfile.NamedTemporaryFile(
                mode="w", suffix=".yaml", delete=False, encoding="utf-8"
            ) as f:
                f.write(content)
                tmp_path = f.name
            config_path = tmp_path
            project = req.config_name

        elif req.config_path:
            p = Path(req.config_path)
            if not p.exists():
                raise HTTPException(status_code=404, detail=f"File not found: {req.config_path}")
            config_path = req.config_path
            import yaml
            try:
                data = yaml.safe_load(p.read_text(encoding="utf-8"))
                project = data.get("name", p.stem)
            except Exception:
                project = p.stem
        else:
            raise HTTPException(status_code=400, detail="Provide config_path, config_name, or inline_config")

        run_id = jobs.start(config_path, project, db_path=str(db_path))
        return {"run_id": run_id, "project": project, "status": "running"}

    @app.delete("/api/jobs/{run_id}")
    def cancel_job(run_id: str):
        ok = jobs.cancel(run_id)
        if not ok:
            raise HTTPException(status_code=409, detail="Job not cancellable (not running or not found)")
        return {"cancelled": True}

    @app.get("/api/jobs/{run_id}/logs/stream")
    async def stream_logs(run_id: str):
        """
        SSE endpoint — streams live log lines for a running job.
        If the job is already finished, streams its buffered log then closes.
        """
        job = jobs.get_job(run_id)
        if not job:
            raise HTTPException(status_code=404, detail=f"Job '{run_id}' not found")

        async def event_stream() -> AsyncGenerator[str, None]:
            # First, emit all buffered lines so the client catches up
            for line in list(job.log_lines):
                yield f"data: {line}\n\n"

            if job.status not in ("running", "queued"):
                yield "data: __DONE__\n\n"
                return

            # Subscribe to live updates
            q = job.subscribe()
            try:
                while True:
                    try:
                        line = await asyncio.wait_for(q.get(), timeout=30.0)
                    except asyncio.TimeoutError:
                        yield "data: __PING__\n\n"
                        continue
                    if line == "__DONE__":
                        yield "data: __DONE__\n\n"
                        break
                    yield f"data: {line}\n\n"
            finally:
                job.unsubscribe(q)

        return StreamingResponse(
            event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "X-Accel-Buffering": "no",
            },
        )

    # ---------------------------------------------------------------- configs

    @app.get("/api/configs")
    def list_configs():
        return cfgs.list_configs()

    @app.get("/api/configs/{name}")
    def get_config(name: str):
        try:
            content = cfgs.read_config(name)
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail=f"Config '{name}' not found")
        return {"name": name, "content": content}

    @app.post("/api/configs", status_code=201)
    def create_config(req: WriteConfigRequest):
        try:
            path = cfgs.write_config(req.name, req.content)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc))
        return {"name": req.name, "path": path}

    @app.put("/api/configs/{name}")
    def update_config(name: str, req: WriteConfigRequest):
        try:
            path = cfgs.write_config(name, req.content)
        except ValueError as exc:
            raise HTTPException(status_code=422, detail=str(exc))
        return {"name": name, "path": path}

    @app.delete("/api/configs/{name}", status_code=204)
    def delete_config(name: str):
        try:
            cfgs.delete_config(name)
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail=f"Config '{name}' not found")

    @app.post("/api/configs/validate")
    def validate_config(req: ValidateConfigRequest):
        try:
            result = cfgs.validate_config(req.content)
        except ValueError as exc:
            return {"valid": False, "error": str(exc)}
        return result

    # ---------------------------------------------------------------- files

    @app.get("/api/runs/{run_id}/files")
    def list_run_files(run_id: str):
        """Return all output files recorded for this run, with size and existence."""
        import json as _json
        run = store.get_run(run_id)
        if not run:
            raise HTTPException(status_code=404, detail=f"Run '{run_id}' not found")
        files = []
        for step in run.get("steps", []):
            raw = step.get("output_files") or "[]"
            try:
                paths = _json.loads(raw) if isinstance(raw, str) else raw
            except Exception:
                paths = []
            for p in paths:
                path_obj = Path(p)
                # Always use forward slashes so paths are safe in JS string literals
                fwd_path = path_obj.as_posix()
                files.append({
                    "path": fwd_path,
                    "filename": path_obj.name,
                    "step": step["step"],
                    "exists": path_obj.exists(),
                    "size_bytes": path_obj.stat().st_size if path_obj.exists() else 0,
                    "ext": path_obj.suffix.lower(),
                })
        return files

    @app.get("/api/files/preview")
    def preview_file(path: str, limit: int = 200):
        """
        Preview a file's contents in the browser.
        JSON files: returns metadata + first `limit` items.
        Excel files: returns column names + first `limit` rows.
        Other files: returns first 4 KB as plain text.
        """
        import json as _json
        p = Path(path)
        if not p.exists():
            raise HTTPException(status_code=404, detail="File not found")
        ext = p.suffix.lower()

        if ext == ".json":
            try:
                data = _json.loads(p.read_text(encoding="utf-8"))
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f"Could not parse JSON: {exc}")
            if isinstance(data, dict) and "items" in data:
                items = data["items"][:limit]
                return {
                    "type": "json",
                    "meta": {k: v for k, v in data.items() if k != "items"},
                    "columns": list(items[0].keys()) if items else [],
                    "rows": items,
                    "total": len(data["items"]),
                    "showing": len(items),
                }
            if isinstance(data, list):
                items = data[:limit]
                return {
                    "type": "json",
                    "meta": {},
                    "columns": list(items[0].keys()) if items else [],
                    "rows": items,
                    "total": len(data),
                    "showing": len(items),
                }
            return {"type": "json_object", "data": data}

        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        sheets.append({"sheet": sheet_name, "columns": [], "rows": [], "total": 0})
                        continue
                    columns = [str(c) if c is not None else "" for c in rows[0]]
                    data_rows = []
                    for row in rows[1: limit + 1]:
                        data_rows.append([str(c) if c is not None else "" for c in row])
                    sheets.append({
                        "sheet": sheet_name,
                        "columns": columns,
                        "rows": data_rows,
                        "total": ws.max_row - 1 if ws.max_row else 0,
                        "showing": len(data_rows),
                    })
                wb.close()
                return {"type": "excel", "sheets": sheets}
            except ImportError:
                raise HTTPException(status_code=501, detail="openpyxl not installed")
            except Exception as exc:
                raise HTTPException(status_code=422, detail=f"Could not read Excel: {exc}")

        # Fallback: plain text preview
        try:
            text = p.read_bytes()[:4096].decode("utf-8", errors="replace")
            return {"type": "text", "content": text}
        except Exception as exc:
            raise HTTPException(status_code=422, detail=str(exc))

    @app.get("/api/files/download")
    def download_file(path: str):
        """Serve a file as a download."""
        p = Path(path)
        if not p.exists():
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(str(p), filename=p.name)

    @app.delete("/api/files", status_code=204)
    def delete_file(path: str):
        """Permanently delete a file from disk."""
        p = Path(path)
        if not p.exists():
            return  # already gone — idempotent
        try:
            p.unlink()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Could not delete file: {exc}")

    # ---------------------------------------------------------------- health

    @app.get("/health")
    def health():
        return {"status": "ok", "db": str(db_path), "configs": str(configs_dir)}

    # ---------------------------------------------------------------- static dashboard

    static_dir = Path(__file__).parent / "static"
    static_dir.mkdir(exist_ok=True)

    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/")
    def dashboard():
        index = static_dir / "index.html"
        if index.exists():
            return FileResponse(str(index))
        return {"message": "ScraperKit API", "docs": "/docs"}

    # Catch-all for SPA routing
    @app.get("/{full_path:path}")
    def spa_fallback(full_path: str):
        if full_path.startswith("api/") or full_path == "health":
            raise HTTPException(status_code=404)
        index = static_dir / "index.html"
        if index.exists():
            return FileResponse(str(index))
        raise HTTPException(status_code=404)

    return app


# Module-level app instance for uvicorn/gunicorn
app = create_app(
    db_path=os.environ.get("SCRAPERKIT_DB", "scraperkit.db"),
    configs_dir=os.environ.get("SCRAPERKIT_CONFIGS", "configs"),
)
